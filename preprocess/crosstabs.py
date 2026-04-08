"""Motor de tabulaciones cruzadas basado en pandas.

Toma un :class:`preprocess.tab_plan.TabPlan` y un
:class:`preprocess.responses_reader.ResponsesData`, y produce un xlsx
con una hoja por cada ``CrossSpec``. El xlsx resultante alimenta
directamente al motor de generacion de PPT de la Etapa 1.

Reglas principales:

- ``aggregate=count`` + sin ``values`` usa ``pd.crosstab``.
- ``aggregate in {sum,mean,median,min,max}`` + ``values`` usa
  ``pivot_table`` con la funcion correspondiente.
- ``percentage`` permite normalizar por fila, columna o total.
- ``filter`` aplica ``DataFrame.query(expr)`` antes de tabular.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, Union

import pandas as pd
from openpyxl import Workbook

from preprocess.responses_reader import ResponsesData
from preprocess.tab_plan import (
    CrossSpec,
    TabPlan,
    _safe_sheet_name,
)


class CrosstabsError(Exception):
    """Error base del motor de tabulaciones cruzadas."""


@dataclass
class CrosstabsResult:
    """Resultado de ejecutar un ``TabPlan`` sobre una ``ResponsesData``."""

    output_path: Path
    tables: dict[str, pd.DataFrame] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

    @property
    def n_tables(self) -> int:
        return len(self.tables)

    def summary(self) -> str:
        lines = [
            f"Archivo generado: {self.output_path}",
            f"Tablas producidas: {self.n_tables}",
        ]
        for name, df in self.tables.items():
            lines.append(
                f"  - {name}: {df.shape[0]} filas x {df.shape[1]} cols"
            )
        if self.warnings:
            lines.append("Warnings:")
            lines.extend(f"  - {w}" for w in self.warnings)
        return "\n".join(lines)


# ---------------------------------------------------------------------- #
# Core                                                                    #
# ---------------------------------------------------------------------- #


def _apply_percentage(df: pd.DataFrame, mode: str) -> pd.DataFrame:
    if mode == "none":
        return df
    numeric = df.select_dtypes(include="number")
    if numeric.empty:
        return df

    if mode == "row":
        row_sums = numeric.sum(axis=1).replace(0, pd.NA)
        normalized = numeric.div(row_sums, axis=0) * 100
    elif mode == "column":
        col_sums = numeric.sum(axis=0).replace(0, pd.NA)
        normalized = numeric.div(col_sums, axis=1) * 100
    elif mode == "total":
        total = numeric.values.sum()
        normalized = (numeric / total) * 100 if total else numeric
    else:  # pragma: no cover
        return df

    return normalized.round(1)


def _compute_cross(df: pd.DataFrame, cross: CrossSpec) -> pd.DataFrame:
    # Aplicar filtro previo si existe.
    if cross.filter:
        try:
            df = df.query(cross.filter)
        except Exception as exc:
            raise CrosstabsError(
                f"Cross '{cross.name}': filtro invalido ({cross.filter!r}): {exc}"
            ) from exc

    # Validar que las columnas existan.
    all_cols = set(cross.rows) | set(cross.columns)
    if cross.values:
        all_cols.add(cross.values)
    missing = [c for c in all_cols if c not in df.columns]
    if missing:
        raise CrosstabsError(
            f"Cross '{cross.name}': columnas inexistentes en los datos: {missing}"
        )

    if cross.aggregate == "count" and cross.values is None:
        # Uso directo de pd.crosstab.
        if cross.columns:
            row_vals = [df[c] for c in cross.rows]
            col_vals = [df[c] for c in cross.columns]
            table = pd.crosstab(
                index=row_vals if len(row_vals) > 1 else row_vals[0],
                columns=col_vals if len(col_vals) > 1 else col_vals[0],
            )
        else:
            table = df.groupby(cross.rows).size().to_frame("count")
    else:
        # pivot_table con la funcion de agregacion.
        agg = cross.aggregate if cross.aggregate != "count" else "count"
        values_col = cross.values or cross.rows[0]
        table = pd.pivot_table(
            df,
            index=cross.rows,
            columns=cross.columns or None,
            values=values_col,
            aggfunc=agg,
            fill_value=0,
        )

    table = _apply_percentage(table, cross.percentage)

    # Aplanar MultiIndex a formato tabular simple.
    if isinstance(table.columns, pd.MultiIndex):
        table.columns = [
            "_".join(str(x) for x in col if x != "")
            for col in table.columns.values
        ]
    table = table.reset_index()

    # Rellenar NaN/Inf en columnas numericas para que el xlsx sea serializable
    # y para que el motor de PPT (Etapa 1) no explote al embeberlo.
    for col in table.select_dtypes(include="number").columns:
        table[col] = table[col].replace([float("inf"), float("-inf")], 0).fillna(0)

    # Nombre de la primera columna mas informativo.
    if not cross.columns and cross.aggregate == "count" and "count" in table.columns:
        pass  # ya esta ok
    return table


def run_tab_plan(
    plan: TabPlan,
    data: ResponsesData,
    output_path: Union[str, Path],
) -> CrosstabsResult:
    """Ejecuta un ``TabPlan`` sobre ``data`` y escribe un xlsx.

    Parameters
    ----------
    plan:
        Tab Plan a ejecutar.
    data:
        Datos de respuestas cargados por ``ResponsesReader``.
    output_path:
        Ruta del xlsx de salida (se crea el directorio si no existe).
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    result = CrosstabsResult(output_path=output_path)
    wb = Workbook()
    wb.remove(wb.active)

    for cross in plan.crosses:
        try:
            table = _compute_cross(data.df, cross)
        except CrosstabsError as exc:
            result.warnings.append(str(exc))
            continue

        sheet_name = _safe_sheet_name(cross.name)
        ws = wb.create_sheet(sheet_name)
        ws.append([str(c) for c in table.columns])
        for row in table.itertuples(index=False, name=None):
            ws.append(list(row))

        result.tables[sheet_name] = table

    if not result.tables:
        # Por lo menos dejamos una hoja vacia para no invalidar el xlsx.
        wb.create_sheet("empty")

    wb.save(str(output_path))
    return result


__all__ = ["CrosstabsError", "CrosstabsResult", "run_tab_plan"]
