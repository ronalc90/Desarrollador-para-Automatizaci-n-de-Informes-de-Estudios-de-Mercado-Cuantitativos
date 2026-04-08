"""Tests del pipeline de procesamiento previo (Etapa 3)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from preprocess.crosstabs import CrosstabsError, run_tab_plan
from preprocess.llm_interpreter import (
    MockLLMInterpreter,
    interpret_tab_plan_text,
)
from preprocess.responses_reader import (
    ResponsesReader,
    ResponsesReaderError,
    UnsupportedFormatError,
)
from preprocess.tab_plan import (
    CrossSpec,
    TabPlan,
    TabPlanError,
    load_tab_plan_excel,
    load_tab_plan_yaml,
)


# ---------------------------------------------------------------------- #
# ResponsesReader                                                         #
# ---------------------------------------------------------------------- #


def test_responses_reader_loads_csv(sample_responses_csv: Path) -> None:
    data = ResponsesReader.load(sample_responses_csv)
    assert data.format == "csv"
    assert data.n_rows == 15
    assert set(data.columns) >= {
        "respondent_id",
        "segmento",
        "edad",
        "satisfaccion",
        "nps",
        "gasto",
    }


def test_responses_reader_loads_xlsx(sample_xlsx: Path) -> None:
    data = ResponsesReader.load(sample_xlsx)
    assert data.format == "xlsx"
    assert data.n_rows > 0


def test_responses_reader_unknown_format(tmp_path: Path) -> None:
    bad = tmp_path / "file.unknown"
    bad.write_text("hola")
    with pytest.raises(UnsupportedFormatError):
        ResponsesReader.load(bad)


def test_responses_reader_missing_file(tmp_path: Path) -> None:
    with pytest.raises(ResponsesReaderError):
        ResponsesReader.load(tmp_path / "nope.csv")


def test_describe_includes_columns(sample_responses_csv: Path) -> None:
    data = ResponsesReader.load(sample_responses_csv)
    text = data.describe()
    assert "segmento" in text
    assert "Filas: 15" in text


# ---------------------------------------------------------------------- #
# TabPlan                                                                 #
# ---------------------------------------------------------------------- #


def test_load_tab_plan_yaml(sample_tab_plan_yaml: Path) -> None:
    plan = load_tab_plan_yaml(sample_tab_plan_yaml)
    assert isinstance(plan, TabPlan)
    assert len(plan.crosses) == 3
    assert plan.crosses[0].name == "Satisfaccion por segmento"
    assert plan.crosses[1].aggregate == "mean"
    assert plan.crosses[1].values == "nps"


def test_load_tab_plan_xlsx(sample_tab_plan_xlsx: Path) -> None:
    plan = load_tab_plan_excel(sample_tab_plan_xlsx)
    assert len(plan.crosses) == 2
    names = [c.name for c in plan.crosses]
    assert "Satisfaccion por segmento" in names


def test_tab_plan_rejects_duplicate_names() -> None:
    with pytest.raises(TabPlanError):
        TabPlan.from_dict(
            {
                "crosses": [
                    {"name": "X", "rows": ["a"], "columns": ["b"]},
                    {"name": "X", "rows": ["c"], "columns": ["d"]},
                ]
            }
        )


def test_tab_plan_rejects_empty_rows() -> None:
    with pytest.raises(TabPlanError):
        TabPlan.from_dict(
            {
                "crosses": [
                    {"name": "X", "rows": [], "columns": ["b"]},
                ]
            }
        )


def test_tab_plan_rejects_bad_aggregate() -> None:
    with pytest.raises(TabPlanError):
        TabPlan.from_dict(
            {
                "crosses": [
                    {
                        "name": "X",
                        "rows": ["a"],
                        "columns": ["b"],
                        "aggregate": "pepito",
                    },
                ]
            }
        )


def test_cross_spec_requires_values_for_mean() -> None:
    with pytest.raises(TabPlanError):
        cross = CrossSpec(
            name="X", rows=["a"], columns=["b"], aggregate="mean"
        )
        cross.validate()


# ---------------------------------------------------------------------- #
# Crosstabs                                                               #
# ---------------------------------------------------------------------- #


def test_run_tab_plan_produces_xlsx(
    tmp_path: Path,
    sample_responses_csv: Path,
    sample_tab_plan_yaml: Path,
) -> None:
    data = ResponsesReader.load(sample_responses_csv)
    plan = load_tab_plan_yaml(sample_tab_plan_yaml)

    out = tmp_path / "crosstabs.xlsx"
    result = run_tab_plan(plan, data, out)

    assert result.output_path == out
    assert out.exists()
    assert result.n_tables == 3
    assert not result.warnings

    # El xlsx debe tener las 3 hojas.
    wb = load_workbook(str(out))
    assert len(wb.sheetnames) == 3
    assert any("Satisfaccion" in n for n in wb.sheetnames)

    # La tabla "NPS promedio por segmento" debe tener 3 segmentos.
    for name in wb.sheetnames:
        if "NPS" in name:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            # header + 3 segmentos (Jovenes, Adultos, Mayores)
            assert len(rows) == 4
            # La primera columna es 'segmento'
            header = rows[0]
            assert header[0] == "segmento"
            break
    else:
        pytest.fail("No se encontro la hoja de NPS")


def test_run_tab_plan_with_filter(
    tmp_path: Path, sample_responses_csv: Path
) -> None:
    data = ResponsesReader.load(sample_responses_csv)
    plan = TabPlan.from_dict(
        {
            "crosses": [
                {
                    "name": "Solo jovenes",
                    "rows": ["satisfaccion"],
                    "columns": ["segmento"],
                    "filter": "segmento == 'Jovenes'",
                }
            ]
        }
    )
    result = run_tab_plan(plan, data, tmp_path / "filtered.xlsx")
    assert result.n_tables == 1
    # Despues del filtro solo hay un segmento.
    table = next(iter(result.tables.values()))
    # La primera columna es "satisfaccion" + una columna con los Jovenes.
    assert "Jovenes" in str(list(table.columns))


def test_run_tab_plan_missing_column(
    tmp_path: Path, sample_responses_csv: Path
) -> None:
    data = ResponsesReader.load(sample_responses_csv)
    plan = TabPlan.from_dict(
        {
            "crosses": [
                {
                    "name": "cross invalido",
                    "rows": ["columna_inexistente"],
                    "columns": ["segmento"],
                }
            ]
        }
    )
    result = run_tab_plan(plan, data, tmp_path / "out.xlsx")
    # No debe abortar; guarda el warning.
    assert result.warnings
    assert "columna_inexistente" in result.warnings[0]


# ---------------------------------------------------------------------- #
# LLM Interpreter (mock)                                                  #
# ---------------------------------------------------------------------- #


def test_mock_interpreter_basic() -> None:
    text = """
    1. Cruzar satisfaccion por segmento
    2. nps por edad
    """
    plan = interpret_tab_plan_text(text)
    assert len(plan.crosses) == 2
    assert plan.crosses[0].rows == ["satisfaccion"]
    assert plan.crosses[0].columns == ["segmento"]
    assert plan.crosses[1].rows == ["nps"]


def test_mock_interpreter_deduplicates() -> None:
    text = "satisfaccion por segmento\nsatisfaccion por segmento"
    plan = interpret_tab_plan_text(text)
    assert len(plan.crosses) == 1


def test_mock_interpreter_empty_raises() -> None:
    with pytest.raises(TabPlanError):
        interpret_tab_plan_text("")


def test_mock_interpreter_no_patterns() -> None:
    with pytest.raises(TabPlanError):
        interpret_tab_plan_text("Quiero un informe lindo y muy bueno")


def test_interpret_then_run_end_to_end(
    tmp_path: Path, sample_responses_csv: Path
) -> None:
    """Flujo Etapa 3 completo: texto libre -> TabPlan -> xlsx crosstabs."""
    text = """
    satisfaccion por segmento
    nps por segmento
    """
    plan = interpret_tab_plan_text(text)
    data = ResponsesReader.load(sample_responses_csv)
    out = tmp_path / "llm_out.xlsx"
    result = run_tab_plan(plan, data, out)
    assert out.exists()
    # Deberia producir las 2 tablas a partir del texto libre.
    assert result.n_tables == 2
