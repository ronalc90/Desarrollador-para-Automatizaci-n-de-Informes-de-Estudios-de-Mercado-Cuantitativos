"""Tests del backend FastAPI (Etapa 2)."""

from __future__ import annotations

import io
import zipfile
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from web.app import create_app


@pytest.fixture
def client(
    tmp_path: Path,
    sample_mapping_multi: Path,
) -> TestClient:
    app = create_app(
        workdir=tmp_path / "workdir",
        default_mapping=sample_mapping_multi,
    )
    return TestClient(app)


def test_health(client: TestClient) -> None:
    resp = client.get("/api/health")
    assert resp.status_code == 200
    body = resp.json()
    assert body["status"] == "ok"
    assert "workdir" in body


def test_ui_root_returns_html(client: TestClient) -> None:
    resp = client.get("/")
    assert resp.status_code == 200
    assert "text/html" in resp.headers["content-type"]
    assert "PPT Engine" in resp.text


def test_ui_app_has_form(client: TestClient) -> None:
    resp = client.get("/app")
    assert resp.status_code == 200
    assert "Generar PPT" in resp.text


def test_inspect_endpoint(client: TestClient, sample_pptx: Path) -> None:
    with sample_pptx.open("rb") as fh:
        resp = client.post(
            "/api/inspect",
            files={
                "template": (
                    "template.pptx",
                    fh,
                    "application/vnd.openxmlformats-officedocument.presentationml.presentation",
                )
            },
        )
    assert resp.status_code == 200
    body = resp.json()
    assert body["total_slides"] == 3
    assert body["total_charts"] == 3
    assert any(
        c["shape_name"] == "Grafico satisfaccion"
        for s in body["slides"]
        for c in s["charts"]
    )
    assert "slide_index: 2" in body["mapping_stub"]


def test_inspect_rejects_non_pptx(client: TestClient) -> None:
    resp = client.post(
        "/api/inspect",
        files={"template": ("fake.txt", b"hello", "text/plain")},
    )
    assert resp.status_code == 400


def test_validate_endpoint_ok(
    client: TestClient,
    sample_pptx: Path,
    sample_xlsx: Path,
    sample_mapping_multi: Path,
) -> None:
    with sample_pptx.open("rb") as t, sample_xlsx.open("rb") as d, sample_mapping_multi.open(
        "rb"
    ) as m:
        resp = client.post(
            "/api/validate",
            files={
                "template": ("template.pptx", t, "application/octet-stream"),
                "data": ("data.xlsx", d, "application/octet-stream"),
                "mapping": ("mapping.yaml", m, "text/yaml"),
            },
        )
    assert resp.status_code == 200
    body = resp.json()
    assert body["ok"] is True
    assert body["errors"] == []


def test_validate_endpoint_with_bad_mapping(
    client: TestClient,
    sample_pptx: Path,
    sample_xlsx: Path,
) -> None:
    bad_mapping = (
        "slides:\n"
        "  - slide_index: 999\n"
        "    charts:\n"
        "      - chart_name: X\n"
        "        excel_sheet: hoja_inexistente\n"
        "        data_range: A1:B2\n"
    ).encode("utf-8")

    with sample_pptx.open("rb") as t, sample_xlsx.open("rb") as d:
        resp = client.post(
            "/api/validate",
            files={
                "template": ("template.pptx", t, "application/octet-stream"),
                "data": ("data.xlsx", d, "application/octet-stream"),
                "mapping": ("mapping.yaml", io.BytesIO(bad_mapping), "text/yaml"),
            },
        )
    assert resp.status_code == 200
    body = resp.json()
    assert body["ok"] is False
    assert body["errors"]


def test_create_job_and_download(
    client: TestClient,
    sample_pptx: Path,
    sample_xlsx: Path,
    sample_mapping_multi: Path,
) -> None:
    with sample_pptx.open("rb") as t, sample_xlsx.open("rb") as d, sample_mapping_multi.open(
        "rb"
    ) as m:
        resp = client.post(
            "/api/jobs",
            files={
                "template": ("template.pptx", t, "application/octet-stream"),
                "data": ("data.xlsx", d, "application/octet-stream"),
                "mapping": ("mapping.yaml", m, "text/yaml"),
            },
        )
    assert resp.status_code == 200
    body = resp.json()
    assert body["status"] == "success"
    assert body["charts_updated"] == 3
    assert body["charts_failed"] == 0
    assert body["output_available"] is True

    job_id = body["job_id"]

    # Consulta del estado.
    state_resp = client.get(f"/api/jobs/{job_id}")
    assert state_resp.status_code == 200
    assert state_resp.json()["status"] == "success"

    # Descarga del output.
    dl_resp = client.get(f"/api/jobs/{job_id}/download")
    assert dl_resp.status_code == 200
    assert (
        "presentationml" in dl_resp.headers["content-type"]
        or "officedocument" in dl_resp.headers["content-type"]
    )

    # El archivo descargado debe ser un .pptx valido con el xlsx embebido.
    pptx_bytes = dl_resp.content
    with zipfile.ZipFile(io.BytesIO(pptx_bytes)) as zf:
        embedded = [n for n in zf.namelist() if n.startswith("ppt/embeddings/")]
        assert embedded, "el pptx descargado no tiene xlsx embebido"
        with zf.open(embedded[0]) as eh:
            wb = load_workbook(io.BytesIO(eh.read()))
        assert wb.active.max_row > 1


def test_create_job_with_default_mapping(
    client: TestClient,
    sample_pptx: Path,
    sample_xlsx: Path,
) -> None:
    """Si no se envia mapping en el request se usa el default del server."""
    with sample_pptx.open("rb") as t, sample_xlsx.open("rb") as d:
        resp = client.post(
            "/api/jobs",
            files={
                "template": ("template.pptx", t, "application/octet-stream"),
                "data": ("data.xlsx", d, "application/octet-stream"),
            },
        )
    assert resp.status_code == 200
    assert resp.json()["status"] == "success"


def test_job_not_found(client: TestClient) -> None:
    resp = client.get("/api/jobs/nonexistent")
    assert resp.status_code == 404
    resp2 = client.get("/api/jobs/nonexistent/download")
    assert resp2.status_code == 404


def test_create_job_validation_failure(
    client: TestClient,
    sample_pptx: Path,
    sample_xlsx: Path,
) -> None:
    bad_mapping = (
        "slides:\n"
        "  - slide_index: 999\n"
        "    charts:\n"
        "      - chart_name: X\n"
        "        excel_sheet: hoja_inexistente\n"
        "        data_range: A1:B2\n"
    ).encode("utf-8")

    with sample_pptx.open("rb") as t, sample_xlsx.open("rb") as d:
        resp = client.post(
            "/api/jobs",
            files={
                "template": ("template.pptx", t, "application/octet-stream"),
                "data": ("data.xlsx", d, "application/octet-stream"),
                "mapping": ("mapping.yaml", io.BytesIO(bad_mapping), "text/yaml"),
            },
        )
    assert resp.status_code == 400
    body = resp.json()
    assert body["status"] == "error"
    assert body["errors"]
