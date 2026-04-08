"""Lectura y parsing del archivo Excel de datos.

Expone una API simple para obtener tablas del Excel por hoja + rango o
por nombre de tabla, devolviendo siempre un ``pandas.DataFrame``.

Ejemplo de uso::

    reader = ExcelReader("data/estudio.xlsx")
    df = reader.get_table("P1_satisfaccion", "A1:E6")
    reader.close()

O como context manager::

    with ExcelReader("data/estudio.xlsx") as reader:
        df = reader.get_table("P1_satisfaccion", "A1:E6")
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook import Workbook


class ExcelReaderError(Exception):
    """Error base para el módulo ``excel_reader``."""


class ExcelFileNotFoundError(ExcelReaderError):
    """El archivo Excel indicado no existe."""


class SheetNotFoundError(ExcelReaderError):
    """La hoja solicitada no existe en el workbook."""


class RangeNotFoundError(ExcelReaderError):
    """El rango solicitado está vacío o fuera de los límites de la hoja."""


class InvalidRangeError(ExcelReaderError):
    """El rango indicado no tiene formato A1 válido."""


class TableNotFoundError(ExcelReaderError):
    """No existe una tabla con ese nombre en la hoja indicada."""


class ExcelReader:
    """Lector de archivos Excel para el motor de PPT.

    Parameters
    ----------
    path:
        Ruta al archivo ``.xlsx``.
    data_only:
        Si ``True`` (por defecto) las celdas con fórmulas devuelven su
        valor calculado en lugar de la fórmula en texto.
    """

    def __init__(self, path: Union[str, Path], data_only: bool = True) -> None:
        self.path = Path(path)
        if not self.path.exists():
            raise ExcelFileNotFoundError(
                f"Archivo Excel no encontrado: {self.path}"
            )
        if self.path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ExcelReaderError(
                f"Extensión no soportada ({self.path.suffix}). "
                "Se esperaba .xlsx o .xlsm."
            )

        try:
            self._wb: Workbook = load_workbook(
                filename=str(self.path),
                data_only=data_only,
                read_only=False,
            )
        except Exception as exc:  # openpyxl lanza varias excepciones
            raise ExcelReaderError(
                f"No se pudo abrir el Excel {self.path}: {exc}"
            ) from exc

    # ------------------------------------------------------------------ #
    # API pública                                                        #
    # ------------------------------------------------------------------ #

    @property
    def sheet_names(self) -> list[str]:
        """Lista de nombres de hoja disponibles."""
        return list(self._wb.sheetnames)

    def has_sheet(self, sheet_name: str) -> bool:
        """Devuelve ``True`` si la hoja existe (comparación exacta)."""
        return sheet_name in self._wb.sheetnames

    def get_table(
        self,
        sheet_name: str,
        data_range: Optional[str] = None,
        *,
        table_id: Optional[str] = None,
        header: bool = True,
    ) -> pd.DataFrame:
        """Obtiene una tabla del Excel como ``DataFrame``.

        Se puede identificar la tabla de dos formas mutuamente excluyentes:

        1. ``data_range``: un rango en notación A1, por ejemplo ``"A1:E6"``.
        2. ``table_id``: el nombre de una tabla Excel (objeto ``Table``
           definido en la hoja).

        Parameters
        ----------
        sheet_name:
            Nombre de la hoja que contiene la tabla.
        data_range:
            Rango en notación A1. Si se omite, debe pasarse ``table_id``.
        table_id:
            Nombre de una ``Table`` de Excel en la hoja indicada.
        header:
            Si ``True`` (por defecto) la primera fila del rango se toma
            como encabezado de las columnas del DataFrame.

        Returns
        -------
        pandas.DataFrame
            DataFrame con los datos de la tabla. El dtype se infiere
            automáticamente desde los valores de las celdas.

        Raises
        ------
        SheetNotFoundError
            Si ``sheet_name`` no existe en el workbook.
        InvalidRangeError
            Si ``data_range`` tiene un formato inválido.
        RangeNotFoundError
            Si el rango está vacío o fuera de límites.
        TableNotFoundError
            Si ``table_id`` no existe en la hoja.
        ValueError
            Si no se pasa ``data_range`` ni ``table_id`` o si se pasan ambos.
        """
        if (data_range is None) == (table_id is None):
            raise ValueError(
                "Debe indicarse exactamente uno de 'data_range' o 'table_id'."
            )

        if not self.has_sheet(sheet_name):
            raise SheetNotFoundError(
                f"La hoja '{sheet_name}' no existe en {self.path.name}. "
                f"Disponibles: {self.sheet_names}"
            )

        ws = self._wb[sheet_name]

        if table_id is not None:
            data_range = self._resolve_table_range(ws, table_id)

        try:
            min_col, min_row, max_col, max_row = range_boundaries(data_range)
        except Exception as exc:
            raise InvalidRangeError(
                f"Rango inválido '{data_range}' en hoja '{sheet_name}': {exc}"
            ) from exc

        rows = list(
            ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
        )

        if not rows or all(all(cell is None for cell in row) for row in rows):
            raise RangeNotFoundError(
                f"El rango '{data_range}' en la hoja '{sheet_name}' "
                "está vacío o fuera de límites."
            )

        if header:
            raw_header = rows[0]
            columns = [
                (f"col_{idx}" if value is None else str(value))
                for idx, value in enumerate(raw_header)
            ]
            data = rows[1:]
        else:
            columns = [f"col_{i}" for i in range(len(rows[0]))]
            data = rows

        df = pd.DataFrame(data, columns=columns)
        return df

    def list_tables(self, sheet_name: str) -> list[str]:
        """Nombres de las tablas Excel definidas en la hoja."""
        if not self.has_sheet(sheet_name):
            raise SheetNotFoundError(
                f"La hoja '{sheet_name}' no existe en {self.path.name}."
            )
        return list(self._wb[sheet_name].tables.keys())

    def close(self) -> None:
        """Cierra el workbook y libera el archivo."""
        try:
            self._wb.close()
        except Exception:
            pass

    # ------------------------------------------------------------------ #
    # Helpers internos                                                   #
    # ------------------------------------------------------------------ #

    @staticmethod
    def _resolve_table_range(ws, table_id: str) -> str:
        tables = ws.tables
        if table_id not in tables:
            raise TableNotFoundError(
                f"No existe la tabla '{table_id}' en la hoja '{ws.title}'. "
                f"Disponibles: {list(tables.keys())}"
            )
        return tables[table_id].ref

    # ------------------------------------------------------------------ #
    # Context manager                                                    #
    # ------------------------------------------------------------------ #

    def __enter__(self) -> "ExcelReader":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()
